#! python3

#Programm ermittelt mithilfe Datensatz von Excel und "best fit"
#(in least-Squares Sense) ein Temperatur Kompensations Polynom
#interpoliert dadurch die Abkühl/Widerstands Kurve für Spulen

#Author     Grabichler
#Date:      06.11.2018
#Company:   Gema-Technik


import numpy 
import matplotlib.pyplot as plt
import openpyxl as py
import os
import pprint

#-----------------------------------------------------------
#importieren der Excel Daten
os.chdir('C:\\Users\\Grabichler\\AppData\\Local\\Programs\\Python\\Python37-32\\wrk\\Gema')
print(os.getcwd())

wb = py.load_workbook('Temperaturkoeffizient-Ermittlung-Polynom-2018-05-04.xlsx', 'r')

ws = wb.sheetnames

sheet = wb['Messwerte 01479']

print(sheet.title)  # Ausgabe von "Messwerte 01479"
#-------------------------------------------------------------------------------------------------




#For Loop, um die gewünschten Daten abzugreifen
#nimmt Werte von 4 Nestern, bildet daraus Mittelwerte
tempDataN1 = []
tempDataN2 = []
tempDataN3 = []
tempDataN4 = []

resDataN1 = []
resDataN2 = []
resDataN3 = []
resDataN4 = []

coefficientDataN1 = []
coefficientDataN2 = []
coefficientDataN3 = []
coefficientDataN4 = []

resistDataN1_20C=[]
resistDataN2_20C = []
resistDataN3_20C = []
resistDataN4_20C = []

tempDataMittelwert = []
resistDataMittelwert = []
coefficientDataMittelwert = []

alphaCoeffCU = 0.00393


#-------------------------------------------------------------------------------------------------
#Daten werden eingelesen
def funcResistanceData(column, row):
    resistData_20C = []
    resistData_20C.insert(7, sheet[column +str(row)].value)
    return resistData_20C

resistDataN1_20C    = funcResistanceData('AF', 7)
resistDataN2_20C    = funcResistanceData('AI', 7)
resistDataN3_20C    = funcResistanceData('AM', 7)
print('----')


for row in range(9, sheet.max_row + 1, 1):
    tempDataN1.insert(row, sheet['AF' + str(row)].value)
    tempDataN2.insert(row, sheet['AI' + str(row)].value)
    tempDataN3.insert(row, sheet['AM' + str(row)].value)
    
    tempDataN1 = [i for i in tempDataN1 if i is not None]
    tempDataN2 = [i for i in tempDataN2 if i is not None]
    tempDataN3 = [i for i in tempDataN3 if i is not None]

    resDataN1.insert(row, sheet['AG' + str(row)].value)
    resDataN2.insert(row, sheet['AJ' + str(row)].value)
    resDataN3.insert(row, sheet['AN' + str(row)].value)
      
    resDataN1 = [i for i in resDataN1 if i is not None]
    resDataN2 = [i for i in resDataN2 if i is not None]
    resDataN3 = [i for i in resDataN3 if i is not None]

for i in range(0, len(tempDataN1), 1):
    tempDataMittelwert.insert(i, numpy.mean([tempDataN1[i], tempDataN2[i], tempDataN3[i]]))
    resistDataMittelwert.insert(i, numpy.mean([resDataN1[i], resDataN2[i], resDataN3[i]]))
    resistDataMittelwert_20C = numpy.mean([resistDataN1_20C, resistDataN2_20C, resistDataN3_20C])
    
    #TemepraturKoeffizientenwerden berechnet aus Temperatur und Widerstand
    #notwendig um die Polynomparameter zu bestimmmen
    coefficientDataMittelwert.insert(i, (resistDataMittelwert[i] - resistDataMittelwert_20C) / ((tempDataMittelwert[i]-20) * resistDataMittelwert_20C))
    
#print(coefficientData)
#print(tempDataMittelwert)


#-------------------------------------------------------------------------------------------------


#for i in range(0, len(xData), 1):
#    coefficientData.insert(i, (yData[i] - resistData) / ((xData[i]-20) * resistData))
    
#xDataLen = (len(xData))
#yDataLen = (len(yData))
#coefficientDataLen = (len(coefficientData))

#DataLen = False
#print(xDataLen)
#print(yDataLen)

#if xDataLen == yDataLen:
#    DataLen = True
#print(DataLen)    
#print(xData)
#print(yData)
#print(coData)

#-------------------------------------------------------------------------------------------------


#-------------------------------------------------------------------------------------------------
#Least-squares fit of a polynomial to data
#coeffs = [1, 2, ]
#coeffs
#if DataLen is True:
coeffs = numpy.polyfit(tempDataMittelwert, coefficientDataMittelwert, 2)
coeffs5 = numpy.polyfit(tempDataMittelwert, coefficientDataMittelwert, 4)
ffit5 = numpy.poly1d(coeffs5)
print(ffit5)
print(coeffs5[0])
print(coeffs5[1])
print(coeffs5[3])
print(coeffs5[4])
calcCoefficientDataPolynom5 = []
for i in range(0, len(tempDataN1), 1):
   calcCoefficientDataPolynom5.insert(i, (coeffs5[0]* tempDataMittelwert[i] ** 4 + (coeffs5[1] * tempDataMittelwert[i] ** 3) + coeffs5[2] * tempDataMittelwert[i] ** 2 + coeffs5[3] * tempDataMittelwert[i] + coeffs5[4]))
#else:
#    print("DatenLänge stimmt nicht überein")
    
#Erzeugt aus Ausgabe [1, 2, 3] 1x^2 + 2x + 3
ffit = numpy.poly1d(coeffs)
#print(ffit)
#-------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------
# Berechnung der Temp Koeffizienten mit den PolynomParameter
calcCoefficientDataPolynom = []

for i in range(0, len(tempDataN1), 1):
   calcCoefficientDataPolynom.insert(i, (coeffs[0]* tempDataMittelwert[i] * tempDataMittelwert[i]) + (coeffs[1] * tempDataMittelwert[i]) + coeffs[2])

#print(calcCoefficientDataPolynom)
#-------------------------------------------------------------------------------------------------

test = []
for i in range(0, len(tempDataN1), 1):
   test.insert(i, (resistDataMittelwert_20C))
#-------------------------------------------------------------------------------------------------
#Berechnung der Nennwiderstände mit Polynom
calculatedResistData = []
for i in range(0, len(tempDataMittelwert), 1):
   calculatedResistData.insert(i, (resistDataMittelwert[i] / (1 + (tempDataMittelwert[i] - 20) * calcCoefficientDataPolynom[i])))

calculatedResistData5 = []
for i in range(0, len(tempDataMittelwert), 1):
   calculatedResistData5.insert(i, (resistDataMittelwert[i] / (1 + (tempDataMittelwert[i] - 20) * calcCoefficientDataPolynom5[i])))                                      

#print(calcCoefficientDataPolynom)                        
#-----------------------------------------------------------


#-----------------------------------------------------------
#ERzeugt Graph wie Matlab und zeigt ihn an

plt.plot(tempDataMittelwert, calculatedResistData5)
plt.plot(tempDataMittelwert, test)
plt.plot(tempDataMittelwert, calculatedResistData)
plt.axis([100,20, 28, 30])
plt.grid()
plt.show()
#-----------------------------------------------------------

#-----------------------------------------------------------
#Exportieren der Daten in eine Excel Tabelle
os.chdir('C:\\Users\\Grabichler\\AppData\\Local\\Programs\\Python\\Python37-32\\wrk\\Gema')
#print(os.getcwd())

print('Writing results in \'Temperaturkoeffizient-Polynom.xlsx\'')

temperaturKoeffizientenPolynom = py.Workbook()
ws1 = temperaturKoeffizientenPolynom.sheetnames

sheet = temperaturKoeffizientenPolynom.active
sheet.title = 'Polynom_Koeffizienten'
print(sheet.title)

sheet['A1'] = 'Polynom Koeffizienten'
sheet['A2'] = 'C0'
sheet['C2'] = 'C1'
sheet['E2'] = 'C2'
sheet['G2'] = 'C3'
sheet['I2'] = 'C4'

sheet['B3'] = 'x^4 + '
sheet['D3'] = 'x^3 + '
sheet['F3'] = 'x^2 + '
sheet['H3'] = 'x^1 + '
sheet['J3'] = 'x^0'

sheet['A3'] = coeffs5[0]
sheet['C3'] = coeffs5[1]
sheet['E3'] = coeffs5[2]
sheet['G3'] = coeffs5[3]
sheet['I3'] = coeffs5[4]


temperaturKoeffizientenPolynom.save('Temperaturkoeffizient-Polynom.xlsx')



#-------------------------------------------------------------------------------------------------

